import os
from celery import shared_task
from celery.exceptions import SoftTimeLimitExceeded
from django.utils import timezone
from django.core.files.base import ContentFile
from tablib import Dataset

from .models import ExportFile, ExportChunk, Respondent
from .resources import RespondentExportResource


from django.core.files.base import File
from tempfile import NamedTemporaryFile
from openpyxl import Workbook


@shared_task(bind=True, soft_time_limit=1800, time_limit=2100)
def export_respondents_task(self, export_file_id):
    try:
        export_file = ExportFile.objects.get(id=export_file_id)
        export_file.status = "processing"
        export_file.save()

        # ресурс + queryset
        resource = RespondentExportResource(
            poll=export_file.poll,
            include_unfinished=export_file.include_unfinished,
        )
        queryset = resource.get_export_queryset(None)
        export_fields = resource.get_export_fields()

        # создаем Excel-файл
        wb = Workbook(write_only=True)  # write_only → не держит всё в памяти
        ws = wb.create_sheet(title="Respondents")

        # заголовки
        headers = [f.column_name for f in export_fields]
        ws.append(headers)

        # данные (итерация по чанкам)
        rows_exported = 0
        for respondent in queryset.iterator(chunk_size=500):
            row = resource.export_resource(respondent)
            ws.append([row.get(f.attribute or f.column_name, "") for f in export_fields])
            rows_exported += 1

        # сохраняем во временный файл
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        poll_id = export_file.poll.id if export_file.poll else "all"
        filename = f"respondents_poll_{poll_id}_{timestamp}.xlsx"

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            export_file.file.save(filename, File(tmp), save=False)

        export_file.filename = filename
        export_file.status = "completed"
        export_file.completed_at = timezone.now()
        export_file.save()

        return {
            "status": "success",
            "export_file_id": export_file_id,
            "rows_exported": rows_exported,
            "filename": filename,
        }

    except ExportFile.DoesNotExist:
        return {"status": "error", "message": f"ExportFile with id {export_file_id} not found"}

    except SoftTimeLimitExceeded:
        try:
            export_file = ExportFile.objects.get(id=export_file_id)
            export_file.status = "failed"
            export_file.error_message = "Task timed out - export took too long"
            export_file.save()
        except ExportFile.DoesNotExist:
            pass
        return {"status": "error", "message": "Task timed out - export took too long"}

    except Exception as e:
        try:
            export_file = ExportFile.objects.get(id=export_file_id)
            export_file.status = "failed"
            export_file.error_message = str(e)
            export_file.save()
        except ExportFile.DoesNotExist:
            pass
        return {"status": "error", "message": str(e)}


@shared_task(bind=True, soft_time_limit=1800, time_limit=2100)  # 30 min soft, 35 min hard
def start_notification_campaign_task(self, campaign_id):
    """
    Основная задача для запуска кампании уведомлений.
    Разбивает пользователей на группы по 100 и запускает дочерние задачи.
    """
    try:
        from .models import NotificationCampaign, Poll
        from apps.users.models import TGUser
        
        campaign = NotificationCampaign.objects.get(id=campaign_id)
        campaign.status = 'processing'
        campaign.save()
        
        # Получаем всех пользователей, которые не прошли опрос по данной теме
        topic = campaign.topic
        users_who_completed = Respondent.objects.filter(
            poll=topic,
            finished_at__isnull=False
        ).values_list('tg_user_id', flat=True).distinct()
        
        # Получаем всех пользователей, которые НЕ прошли опрос по данной теме
        # Исключаем заблокированных пользователей
        all_users = TGUser.objects.filter(is_active=True, blocked_bot=False)
        users_to_notify = all_users.exclude(id__in=users_who_completed)
        
        campaign.total_users = users_to_notify.count()
        campaign.save()
        
        if campaign.total_users == 0:
            campaign.status = 'completed'
            campaign.completed_at = timezone.now()
            campaign.save()
            return {
                'status': 'success',
                'message': 'No users to notify for this topic'
            }
        
        # Разбиваем пользователей на группы по 100
        user_ids = list(users_to_notify.values_list('id', flat=True))
        chunk_size = 100
        user_chunks = [user_ids[i:i + chunk_size] for i in range(0, len(user_ids), chunk_size)]
        
        # Запускаем дочерние задачи последовательно
        for i, chunk in enumerate(user_chunks):
            # Запускаем задачу с задержкой для последовательности
            send_notifications_chunk_task.apply_async(
                args=[campaign_id, chunk, i],
                countdown=i * 2  # 2 секунды между запусками
            )
        
        return {
            'status': 'success',
            'message': f'Started notification campaign for {campaign.total_users} users in {len(user_chunks)} chunks'
        }
        
    except NotificationCampaign.DoesNotExist:
        return {
            'status': 'error',
            'message': f'NotificationCampaign with id {campaign_id} not found'
        }
    except SoftTimeLimitExceeded:
        # Обработка таймаута
        try:
            campaign = NotificationCampaign.objects.get(id=campaign_id)
            campaign.status = 'failed'
            campaign.error_message = 'Task timed out - campaign took too long'
            campaign.save()
        except NotificationCampaign.DoesNotExist:
            pass
        
        return {
            'status': 'error',
            'message': 'Task timed out - campaign took too long'
        }
    except Exception as e:
        # Обновляем статус на "ошибка"
        try:
            campaign = NotificationCampaign.objects.get(id=campaign_id)
            campaign.status = 'failed'
            campaign.error_message = str(e)
            campaign.save()
        except NotificationCampaign.DoesNotExist:
            pass
        
        return {
            'status': 'error',
            'message': str(e)
        }


@shared_task(bind=True, soft_time_limit=300, time_limit=360)  # 5 min soft, 6 min hard
def send_notifications_chunk_task(self, campaign_id, user_ids, chunk_index):
    """
    Отправляет уведомления группе пользователей (до 100 человек).
    Учитывает интервал отправки для избежания блокировки.
    """
    try:
        from .models import NotificationCampaign, Poll
        from apps.users.models import TGUser
        from apps.bot.misc import get_bot_instance
        from django.utils.translation import gettext as _
        from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton
        import time
        
        campaign = NotificationCampaign.objects.get(id=campaign_id)
        topic = campaign.topic
        
        # Получаем пользователей для уведомления
        # В TGUser.id хранится telegram_id (chat_id для бота)
        # Исключаем заблокированных пользователей
        users = TGUser.objects.filter(id__in=user_ids, is_active=True, blocked_bot=False)
        
        sent_count = 0
        failed_count = 0
        
        for i, user in enumerate(users):
            try:
                # Создаем клавиатуру с кнопками
                markup = InlineKeyboardMarkup(inline_keyboard=[
                    [
                        InlineKeyboardButton(text="🔄 Давом этиш", callback_data=f"poll_continue:{topic.uuid}"),
                        InlineKeyboardButton(text="♻️ Қайта бошлаш", callback_data=f"poll_restart:{topic.uuid}")
                    ]
                ])
                
                # Текст уведомления
                message_text = str(_("Сиз сўровномани тўлиқ якунламагансиз. Давом этасизми ёки қайта бошлайсизми?"))
                
                # Отправляем сообщение синхронно
                bot = get_bot_instance()
                # Используем синхронный метод для отправки
                import asyncio
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    loop.run_until_complete(
                        bot.send_message(
                            chat_id=user.id,  # В TGUser.id хранится telegram_id
                            text=message_text,
                            reply_markup=markup
                        )
                    )
                finally:
                    # Правильно закрываем сессию бота
                    loop.run_until_complete(bot.session.close())
                    loop.close()
                
                sent_count += 1
                
                # Обновляем счетчик в кампании
                campaign.sent_users += 1
                campaign.save()
                
                # Пауза между отправками (1 секунда)
                if i < len(users) - 1:  # Не ждем после последнего пользователя
                    time.sleep(1)
                
            except Exception as e:
                failed_count += 1
                # Логируем ошибку, но продолжаем с другими пользователями
                print(f"Failed to send notification to user {user.id}: {e}")
                
                # Проверяем, является ли ошибка связанной с блокировкой бота
                error_message = str(e).lower()
                if any(keyword in error_message for keyword in [
                    'bot was blocked', 'user is deactivated', 'chat not found',
                    'forbidden', 'blocked', 'deactivated', 'bot blocked by user'
                ]):
                    # Помечаем пользователя как заблокировавшего бота
                    user.blocked_bot = True
                    user.save()
                    print(f"Marked user {user.id} as blocked_bot=True due to error: {e}")
                
                continue
        
        # Проверяем, завершена ли вся кампания
        if campaign.sent_users >= campaign.total_users:
            campaign.status = 'completed'
            campaign.completed_at = timezone.now()
            campaign.save()
        
        return {
            'status': 'success',
            'chunk_index': chunk_index,
            'sent_count': sent_count,
            'failed_count': failed_count,
            'message': f'Chunk {chunk_index}: sent {sent_count}, failed {failed_count}'
        }
        
    except NotificationCampaign.DoesNotExist:
        return {
            'status': 'error',
            'message': f'NotificationCampaign with id {campaign_id} not found'
        }
    except SoftTimeLimitExceeded:
        return {
            'status': 'error',
            'message': f'Chunk {chunk_index} timed out'
        }
    except Exception as e:
        return {
            'status': 'error',
            'chunk_index': chunk_index,
            'message': str(e)
        }


@shared_task
def cleanup_old_exports():
    """
    Удаляет старые файлы экспорта (старше 30 дней)
    """
    from datetime import timedelta
    
    cutoff_date = timezone.now() - timedelta(days=30)
    old_exports = ExportFile.objects.filter(
        created_at__lt=cutoff_date,
        status__in=['completed', 'failed']
    )
    
    count = old_exports.count()
    old_exports.delete()
    
    return {
        'status': 'success',
        'deleted_count': count
    }


@shared_task(bind=True, soft_time_limit=1800, time_limit=2100)  # 30 min soft, 35 min hard
def start_broadcast_task(self, broadcast_id):
    """
    Основная задача для запуска рассылки поста.
    Разбивает пользователей на группы по 100 и запускает дочерние задачи.
    """
    try:
        from .models import BroadcastPost
        from apps.users.models import TGUser
        
        broadcast = BroadcastPost.objects.get(id=broadcast_id)
        broadcast.status = 'sending'
        broadcast.started_at = timezone.now()
        broadcast.save()
        
        # Получаем всех активных пользователей, которые не заблокировали бота
        all_users = TGUser.objects.filter(is_active=True, blocked_bot=False)
        broadcast.total_users = all_users.count()
        broadcast.save()
        
        if broadcast.total_users == 0:
            broadcast.status = 'sent'
            broadcast.completed_at = timezone.now()
            broadcast.save()
            return {
                'status': 'success',
                'message': 'No active users to send broadcast to'
            }
        
        # Разбиваем пользователей на группы по 100
        user_ids = list(all_users.values_list('id', flat=True))
        chunk_size = 100
        user_chunks = [user_ids[i:i + chunk_size] for i in range(0, len(user_ids), chunk_size)]
        
        # Запускаем дочерние задачи последовательно
        for i, chunk in enumerate(user_chunks):
            # Запускаем задачу с задержкой для последовательности
            send_broadcast_chunk_task.apply_async(
                args=[broadcast_id, chunk, i],
                countdown=i * 2  # 2 секунды между запусками
            )
        
        return {
            'status': 'success',
            'message': f'Started broadcast for {broadcast.total_users} users in {len(user_chunks)} chunks'
        }
        
    except BroadcastPost.DoesNotExist:
        return {
            'status': 'error',
            'message': f'BroadcastPost with id {broadcast_id} not found'
        }
    except SoftTimeLimitExceeded:
        # Обработка таймаута
        try:
            broadcast = BroadcastPost.objects.get(id=broadcast_id)
            broadcast.status = 'failed'
            broadcast.error_message = 'Task timed out - broadcast took too long'
            broadcast.save()
        except BroadcastPost.DoesNotExist:
            pass
        
        return {
            'status': 'error',
            'message': 'Task timed out - broadcast took too long'
        }
    except Exception as e:
        # Обновляем статус на "ошибка"
        try:
            broadcast = BroadcastPost.objects.get(id=broadcast_id)
            broadcast.status = 'failed'
            broadcast.error_message = str(e)
            broadcast.save()
        except BroadcastPost.DoesNotExist:
            pass
        
        return {
            'status': 'error',
            'message': str(e)
        }


@shared_task(bind=True, soft_time_limit=300, time_limit=360)  # 5 min soft, 6 min hard
def send_broadcast_chunk_task(self, broadcast_id, user_ids, chunk_index):
    """
    Отправляет пост группе пользователей (до 100 человек).
    Учитывает интервал отправки для избежания блокировки.
    """
    try:
        from .models import BroadcastPost
        from apps.users.models import TGUser
        from apps.bot.misc import get_bot_instance
        from aiogram.types import InputFile
        import time
        
        broadcast = BroadcastPost.objects.get(id=broadcast_id)
        
        # Получаем пользователей для рассылки (только активных и не заблокировавших бота)
        users = TGUser.objects.filter(
            id__in=user_ids, 
            is_active=True, 
            blocked_bot=False
        )
        
        sent_count = 0
        failed_count = 0
        
        for i, user in enumerate(users):
            try:
                # Отправляем сообщение синхронно
                bot = get_bot_instance()
                # Используем синхронный метод для отправки
                import asyncio
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    if broadcast.image:
                        # Отправляем с изображением
                        photo = InputFile(broadcast.image.path)
                        loop.run_until_complete(
                            bot.send_photo(
                                chat_id=user.id,
                                photo=photo,
                                caption=f"<b>{broadcast.title}</b>\n\n{broadcast.content}",
                                parse_mode="HTML"
                            )
                        )
                    else:
                        # Отправляем только текст
                        loop.run_until_complete(
                            bot.send_message(
                                chat_id=user.id,
                                text=f"<b>{broadcast.title}</b>\n\n{broadcast.content}",
                                parse_mode="HTML"
                            )
                        )
                finally:
                    # Правильно закрываем сессию бота
                    loop.run_until_complete(bot.session.close())
                    loop.close()
                
                sent_count += 1
                
                # Обновляем счетчик в рассылке
                broadcast.sent_users += 1
                broadcast.save()
                
                # Пауза между отправками (1 секунда)
                if i < len(users) - 1:  # Не ждем после последнего пользователя
                    time.sleep(1)
                
            except Exception as e:
                failed_count += 1
                # Обновляем счетчик ошибок в рассылке
                broadcast.failed_users += 1
                broadcast.save()
                
                # Проверяем, заблокировал ли пользователь бота
                error_message = str(e).lower()
                if any(keyword in error_message for keyword in ['blocked', 'forbidden', 'chat not found']):
                    # Помечаем пользователя как заблокировавшего бота
                    user.blocked_bot = True
                    user.is_active = False
                    user.save()
                    print(f"User {user.id} ({user.fullname}) blocked the bot during broadcast")
                else:
                    print(f"Failed to send broadcast to user {user.id}: {e}")
                continue
        
        # Счетчики ошибок уже обновлены в цикле
        
        # Проверяем, завершена ли вся рассылка
        if broadcast.sent_users + broadcast.failed_users >= broadcast.total_users:
            broadcast.status = 'sent'
            broadcast.completed_at = timezone.now()
            broadcast.save()
        
        return {
            'status': 'success',
            'chunk_index': chunk_index,
            'sent_count': sent_count,
            'failed_count': failed_count,
            'message': f'Chunk {chunk_index}: sent {sent_count}, failed {failed_count}'
        }
        
    except BroadcastPost.DoesNotExist:
        return {
            'status': 'error',
            'message': f'BroadcastPost with id {broadcast_id} not found'
        }
    except SoftTimeLimitExceeded:
        return {
            'status': 'error',
            'message': f'Chunk {chunk_index} timed out'
        }
    except Exception as e:
        return {
            'status': 'error',
            'chunk_index': chunk_index,
            'message': str(e)
        }


@shared_task(bind=True, soft_time_limit=60, time_limit=120)  # 1 min soft, 2 min hard
def send_test_broadcast_task(self, broadcast_id, test_user_id):
    """
    Отправляет тестовый пост конкретному пользователю по Telegram ID.
    Используется для проверки внешнего вида поста перед массовой рассылкой.
    """
    try:
        from .models import BroadcastPost
        from apps.users.models import TGUser
        from apps.bot.misc import get_bot_instance
        from aiogram.types import InputFile
        import asyncio
        
        broadcast = BroadcastPost.objects.get(id=broadcast_id)
        
        # Проверяем, существует ли пользователь с указанным ID
        try:
            test_user = TGUser.objects.get(id=test_user_id)
        except TGUser.DoesNotExist:
            return {
                'status': 'error',
                'message': f'Пользователь с ID {test_user_id} не найден в базе данных'
            }
        
        # Отправляем сообщение
        bot = get_bot_instance()
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        try:
            if broadcast.image:
                # Отправляем с изображением
                photo = InputFile(broadcast.image.path)
                loop.run_until_complete(
                    bot.send_photo(
                        chat_id=test_user_id,
                        photo=photo,
                        caption=f"<b>{broadcast.title}</b>\n\n{broadcast.content}",
                        parse_mode="HTML"
                    )
                )
            else:
                # Отправляем только текст
                loop.run_until_complete(
                    bot.send_message(
                        chat_id=test_user_id,
                        text=f"<b>{broadcast.title}</b>\n\n{broadcast.content}",
                        parse_mode="HTML"
                    )
                )
        finally:
            # Правильно закрываем сессию бота
            loop.run_until_complete(bot.session.close())
            loop.close()
        
        return {
            'status': 'success',
            'message': f'Тестовый пост успешно отправлен пользователю {test_user.fullname} (ID: {test_user_id})'
        }
        
    except BroadcastPost.DoesNotExist:
        return {
            'status': 'error',
            'message': f'BroadcastPost с ID {broadcast_id} не найден'
        }
    except Exception as e:
        return {
            'status': 'error',
            'message': f'Ошибка при отправке тестового поста: {str(e)}'
        }


@shared_task(bind=True, soft_time_limit=1800, time_limit=2100)
def export_respondents_chunked_task(self, export_file_id, chunk_size=1000, max_chunks=10):
    """
    Задача для создания chunked экспорта - разбивает данные на части и запускает параллельные задачи
    """
    try:
        export_file = ExportFile.objects.get(id=export_file_id)
        export_file.status = "processing"
        export_file.save()

        # Получаем ресурс и queryset
        resource = RespondentExportResource(
            poll=export_file.poll,
            include_unfinished=export_file.include_unfinished,
        )
        queryset = resource.get_export_queryset(None)
        total_count = queryset.count()
        
        if total_count == 0:
            export_file.status = "completed"
            export_file.completed_at = timezone.now()
            export_file.save()
            return {"status": "success", "message": "No data to export"}

        # Вычисляем количество chunks
        total_chunks = min(max_chunks, (total_count + chunk_size - 1) // chunk_size)
        
        # Обновляем параметры экспорта
        export_file.is_chunked = True
        export_file.total_chunks = total_chunks
        export_file.chunk_size = chunk_size
        export_file.save()

        # Создаем chunk записи
        chunks = []
        for i in range(total_chunks):
            chunk = ExportChunk.objects.create(
                export_file=export_file,
                chunk_number=i + 1,
                filename=f"{export_file.filename}_part_{i + 1}.xlsx"
            )
            chunks.append(chunk)

        # Запускаем задачи для каждого chunk
        for chunk in chunks:
            export_chunk_task.delay(chunk.id)

        return {
            "status": "success",
            "export_file_id": export_file_id,
            "total_chunks": total_chunks,
            "chunk_size": chunk_size,
            "total_records": total_count
        }

    except ExportFile.DoesNotExist:
        return {"status": "error", "message": f"ExportFile with id {export_file_id} not found"}
    except Exception as e:
        try:
            export_file = ExportFile.objects.get(id=export_file_id)
            export_file.status = "failed"
            export_file.error_message = str(e)
            export_file.save()
        except ExportFile.DoesNotExist:
            pass
        return {"status": "error", "message": str(e)}


@shared_task(bind=True, soft_time_limit=600, time_limit=900)  # 10 min soft, 15 min hard
def export_chunk_task(self, chunk_id):
    """
    Задача для экспорта отдельного chunk
    """
    try:
        chunk = ExportChunk.objects.get(id=chunk_id)
        chunk.status = "processing"
        chunk.save()

        # Получаем основной экспорт
        export_file = chunk.export_file
        resource = RespondentExportResource(
            poll=export_file.poll,
            include_unfinished=export_file.include_unfinished,
        )
        queryset = resource.get_export_queryset(None)
        export_fields = resource.get_export_fields()

        # Вычисляем offset и limit для этого chunk
        offset = (chunk.chunk_number - 1) * export_file.chunk_size
        limit = export_file.chunk_size

        # Получаем данные для этого chunk
        chunk_queryset = queryset[offset:offset + limit]

        # Создаем Excel файл
        wb = Workbook(write_only=True)
        ws = wb.create_sheet(title="Respondents")

        # Заголовки
        headers = [f.column_name for f in export_fields]
        ws.append(headers)

        # Данные
        rows_exported = 0
        for respondent in chunk_queryset:
            row = resource.export_resource(respondent)
            ws.append([row.get(f.attribute or f.column_name, "") for f in export_fields])
            rows_exported += 1

        # Сохраняем файл
        timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{chunk.filename}_{timestamp}.xlsx"

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            chunk.file.save(filename, File(tmp), save=False)

        # Обновляем chunk
        chunk.filename = filename
        chunk.rows_count = rows_exported
        chunk.status = "completed"
        chunk.completed_at = timezone.now()
        chunk.save()

        # Проверяем, завершен ли весь экспорт
        check_export_completion.delay(export_file.id)

        return {
            "status": "success",
            "chunk_id": chunk_id,
            "rows_exported": rows_exported,
            "filename": filename
        }

    except ExportChunk.DoesNotExist:
        return {"status": "error", "message": f"ExportChunk with id {chunk_id} not found"}
    except SoftTimeLimitExceeded:
        try:
            chunk = ExportChunk.objects.get(id=chunk_id)
            chunk.status = "failed"
            chunk.error_message = "Task timed out - chunk export took too long"
            chunk.save()
        except ExportChunk.DoesNotExist:
            pass
        return {"status": "error", "message": "Task timed out - chunk export took too long"}
    except Exception as e:
        try:
            chunk = ExportChunk.objects.get(id=chunk_id)
            chunk.status = "failed"
            chunk.error_message = str(e)
            chunk.save()
        except ExportChunk.DoesNotExist:
            pass
        return {"status": "error", "message": str(e)}


@shared_task(bind=True)
def check_export_completion(self, export_file_id):
    """
    Проверяет, завершен ли весь chunked экспорт
    """
    try:
        export_file = ExportFile.objects.get(id=export_file_id)
        
        # Проверяем статус всех chunks
        completed_chunks = export_file.chunks.filter(status='completed').count()
        failed_chunks = export_file.chunks.filter(status='failed').count()
        total_chunks = export_file.total_chunks
        
        # Обновляем счетчик завершенных chunks
        export_file.completed_chunks = completed_chunks
        export_file.save()
        
        if completed_chunks >= total_chunks:
            # Все chunks завершены
            export_file.status = "completed"
            export_file.completed_at = timezone.now()
            export_file.save()
            
            return {
                "status": "success",
                "message": "All chunks completed",
                "completed_chunks": completed_chunks,
                "total_chunks": total_chunks
            }
        elif failed_chunks > 0 and (completed_chunks + failed_chunks) >= total_chunks:
            # Есть неудачные chunks, но все chunks обработаны
            export_file.status = "failed"
            export_file.error_message = f"Some chunks failed. Completed: {completed_chunks}, Failed: {failed_chunks}"
            export_file.save()
            
            return {
                "status": "error",
                "message": "Some chunks failed",
                "completed_chunks": completed_chunks,
                "failed_chunks": failed_chunks,
                "total_chunks": total_chunks
            }
        else:
            # Экспорт еще продолжается
            return {
                "status": "processing",
                "message": "Export still in progress",
                "completed_chunks": completed_chunks,
                "total_chunks": total_chunks
            }
            
    except ExportFile.DoesNotExist:
        return {"status": "error", "message": f"ExportFile with id {export_file_id} not found"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


@shared_task(bind=True, soft_time_limit=300, time_limit=360)  # 5 min soft, 6 min hard
def send_update_notification_task(self, user_ids, chunk_index, custom_message=None):
    """
    Отправляет уведомление об обновлении группе пользователей (до 100 человек).
    Учитывает интервал отправки для избежания блокировки.
    """
    try:
        from apps.users.models import TGUser
        from apps.bot.misc import get_bot_instance
        import time
        
        # Получаем пользователей для уведомления
        users = TGUser.objects.filter(id__in=user_ids, is_active=True, blocked_bot=False)
        
        # Определяем сообщение для каждого языка
        if custom_message:
            messages = {
                'uz_cyrl': custom_message,
                'uz_latn': custom_message,
                'ru': custom_message
            }
        else:
            messages = {
                'uz_cyrl': (
                    "🎉 <b>Янгиланиш!</b>\n\n"
                    "Ботда янги имкониятлар қўшилди:\n\n"
                    "💰 <b>Пул ишлаш</b> - Сўровномаларни тўлдириб пул ишлаб топинг!\n"
                    "📊 Актив ва якунланган сўровномаларни кўринг\n"
                    "💳 Ишлаб топган пулларингизни чиқаришингиз мумкин\n"
                    "🌐 Тилни ўзгартириш имкони\n\n"
                    "Бошлаш учун /menu буйруғини юборинг!"
                ),
                'uz_latn': (
                    "🎉 <b>Yangilanish!</b>\n\n"
                    "Botda yangi imkoniyatlar qo'shildi:\n\n"
                    "💰 <b>Pul ishlash</b> - So'rovnomalarni to'ldirib pul ishlab toping!\n"
                    "📊 Aktiv va yakunlangan so'rovnomalarni ko'ring\n"
                    "💳 Ishlab topgan pullaringizni chiqarishingiz mumkin\n"
                    "🌐 Tilni o'zgartirish imkoni\n\n"
                    "Boshlash uchun /menu buyrug'ini yuboring!"
                ),
                'ru': (
                    "🎉 <b>Обновление!</b>\n\n"
                    "В боте появились новые возможности:\n\n"
                    "💰 <b>Заработок</b> - Зарабатывайте деньги, заполняя опросы!\n"
                    "📊 Просматривайте активные и завершенные опросы\n"
                    "💳 Выводите заработанные деньги\n"
                    "🌐 Возможность смены языка\n\n"
                    "Чтобы начать, отправьте команду /menu!"
                )
            }
        
        sent_count = 0
        failed_count = 0
        
        for i, user in enumerate(users):
            try:
                # Получаем сообщение на языке пользователя
                message_text = messages.get(user.lang, messages['uz_cyrl'])
                
                # Отправляем сообщение синхронно
                bot = get_bot_instance()
                import asyncio
                loop = asyncio.new_event_loop()
                asyncio.set_event_loop(loop)
                try:
                    loop.run_until_complete(
                        bot.send_message(
                            chat_id=user.id,
                            text=message_text,
                            parse_mode="HTML"
                        )
                    )
                finally:
                    # Правильно закрываем сессию бота
                    loop.run_until_complete(bot.session.close())
                    loop.close()
                
                sent_count += 1
                
                # Пауза между отправками (1 секунда)
                if i < len(users) - 1:  # Не ждем после последнего пользователя
                    time.sleep(1)
                
            except Exception as e:
                failed_count += 1
                # Логируем ошибку, но продолжаем с другими пользователями
                print(f"Failed to send update notification to user {user.id}: {e}")
                
                # Проверяем, является ли ошибка связанной с блокировкой бота
                error_message = str(e).lower()
                if any(keyword in error_message for keyword in [
                    'bot was blocked', 'user is deactivated', 'chat not found',
                    'forbidden', 'blocked', 'deactivated', 'bot blocked by user'
                ]):
                    # Помечаем пользователя как заблокировавшего бота
                    user.blocked_bot = True
                    user.save()
                    print(f"Marked user {user.id} as blocked_bot=True due to error: {e}")
                
                continue
        
        return {
            'status': 'success',
            'chunk_index': chunk_index,
            'sent_count': sent_count,
            'failed_count': failed_count,
            'message': f'Chunk {chunk_index}: sent {sent_count}, failed {failed_count}'
        }
        
    except SoftTimeLimitExceeded:
        return {
            'status': 'error',
            'message': f'Chunk {chunk_index} timed out'
        }
    except Exception as e:
        return {
            'status': 'error',
            'chunk_index': chunk_index,
            'message': str(e)
        }
